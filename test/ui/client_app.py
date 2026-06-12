import re
import sys
import streamlit as st
import pandas as pd
import os
import time
import io
import openpyxl
import json
import sqlite3
import uuid
import datetime
# --- 统一数据库路径 ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# 强制指向 test 目录下的那个有数据的数据库
db_path = os.path.abspath(os.path.join(BASE_DIR, "..", "seo_master.db"))
st.write(f"📢 系统正在使用唯一数据库: {db_path}")
# --- 数据库基础函数 ---
def init_db():
    conn = sqlite3.connect(db_path, check_same_thread=False)
    # 开启 WAL 模式 (特效药)
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA synchronous=NORMAL;")
    cursor = conn.cursor()
    # 1. 用户表：增加密码、到期时间、SessionID
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS users (
            user_id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE,
            password TEXT,
            expiry_date DATE,
            last_session_id TEXT,
            is_active INTEGER DEFAULT 1
        )
    """)
    # 2. 优化历史表：增加 user_id 隔离
    cursor.execute("""
            CREATE TABLE IF NOT EXISTS optimized_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            original_input TEXT,
            optimized_title TEXT,
            platform TEXT,
            char_limit INTEGER,
            opt_mode TEXT,      -- 👈 新增：区分优化模式
            language TEXT,      -- 👈 新增：区分语言
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.commit()
    conn.close()


init_db()  # 每次启动自动检查表结构


# --- 数据管理与超级后台函数 ---
def manage_admin_and_history(conn, current_user_id, is_admin_mode=False):
    st.write("---")
    st.header("🛠️ 数据与账号管理")

    # --- 功能一：个人足迹中心 (所有人可见) ---
    with st.expander("📊 我的优化足迹与清理", expanded=True):
        cursor = conn.cursor()
        # 1. 先只查一个数字（开销极小）
        cursor.execute("SELECT COUNT(*) FROM optimized_history WHERE user_id = ?", (current_user_id,))
        my_count = cursor.fetchone()[0]

        st.info(f"📈 您当前系统中有 **{my_count}** 条优化记录。")

        # 2. 只有点击按钮才查询详情
        if st.button("🔍 点击查询/刷新详细记录", key="btn_query_history"):
            if my_count > 0:
                st.write("📂 您的最近优化记录：")
                history_sample = pd.read_sql_query(f"""
                        SELECT original_input as '原始输入', optimized_title as '优化结果', opt_mode as '模式', timestamp as '时间'
                        FROM optimized_history 
                        WHERE user_id = {current_user_id} 
                        ORDER BY timestamp DESC LIMIT 50
                    """, conn)
                st.dataframe(history_sample, width="stretch")
            else:
                st.warning("目前暂无详细记录。")

        # 3. 清理逻辑（依然放在最下方）
        if my_count > 0:
            st.write("---")
            st.caption("风险操作区")
            confirm_delete = st.checkbox("我已确认：清空后，之前的优化缓存将彻底消失", key="del_confirm_user")
            if st.button("🔥 清空我的专属缓存", disabled=not confirm_delete, type="primary"):
                cursor.execute("DELETE FROM optimized_history WHERE user_id = ?", (current_user_id,))
                conn.commit()
                st.success("清理完成！")
                st.rerun()

    # --- 功能二：超级管理员后台 (仅当密码正确时显示) ---
    if is_admin_mode:
        st.write("---")
        st.subheader("🌟 超级管理员控制台")
        tab1, tab2, tab3 = st.tabs(["账号管理", "数据穿透查询", "系统统计"])

        with tab1:
            st.write("### 用户增删改查")
            with st.form("add_user_form", clear_on_submit=True):
                col1, col2, col3 = st.columns(3)
                new_u = col1.text_input("新账号")
                new_p = col2.text_input("新密码")
                new_d = col3.date_input("到期时间", value=datetime.date.today() + datetime.timedelta(days=30))
                if st.form_submit_button("立即创建账号"):
                    if new_u and new_p:
                        try:
                            conn.execute("INSERT INTO users (username, password, expiry_date) VALUES (?, ?, ?)",
                                         (new_u, new_p, str(new_d)))
                            conn.commit()
                            st.success(f"账号 {new_u} 创建成功")
                        except Exception as e:
                            st.error(f"失败: {e}")

            users_df = pd.read_sql_query("SELECT user_id, username, expiry_date, is_active FROM users", conn)
            st.dataframe(users_df, width="stretch")

            del_user_id = st.number_input("输入要注销的用户 ID", step=1, value=0, key="admin_del_id")
            if st.button("🔥 永久注销该用户", type="secondary"):
                if del_user_id == 1:
                    st.error("不能删除主管理员！")
                elif del_user_id > 0:
                    conn.execute("DELETE FROM users WHERE user_id = ?", (del_user_id,))
                    conn.commit()
                    st.success(f"用户 ID {del_user_id} 已删除")
                    st.rerun()

        with tab2:
            st.write("### 🔍 用户资产穿透查询")
            target_uid = st.number_input("输入要查询的用户 ID", step=1, value=0, key="query_uid")
            if target_uid > 0:
                u_info = conn.execute("SELECT username FROM users WHERE user_id = ?", (target_uid,)).fetchone()
                if u_info:
                    u_count = \
                    conn.execute("SELECT COUNT(*) FROM optimized_history WHERE user_id = ?", (target_uid,)).fetchone()[
                        0]
                    st.metric(f"用户 {u_info[0]} 的总记录数", f"{u_count} 条")

                    if u_count > 0:
                        u_detail = pd.read_sql_query(
                            f"SELECT * FROM optimized_history WHERE user_id = {target_uid} LIMIT 10", conn)
                        st.write("该用户最新的 10 条数据：")
                        st.dataframe(u_detail, width="stretch")

                        if st.button(f"🗑️ 单独清空该用户({u_info[0]})的记录"):
                            conn.execute("DELETE FROM optimized_history WHERE user_id = ?", (target_uid,))
                            conn.commit()
                            st.success(f"{u_info[0]} 的数据已清理")
                    else:
                        st.info(f"用户 {u_info[0]} 目前没有历史记录。")
                else:
                    st.error("找不到该用户 ID")

        with tab3:
            st.write("### 全局系统概览")
            col_a, col_b = st.columns(2)
            total_history = conn.execute("SELECT COUNT(*) FROM optimized_history").fetchone()[0]
            total_users = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
            col_a.metric("系统总缓存条数", total_history)
            col_b.metric("总注册用户数", total_users)

            st.divider()
            st.error("🚨 终极危险区")
            if st.checkbox("我授权执行全库清空"):
                if st.button("清空全库优化历史"):
                    conn.execute("DELETE FROM optimized_history")
                    conn.commit()
                    st.success("全库已清零")

def login_user(username, password):
    conn = sqlite3.connect(db_path, check_same_thread=False)
    cursor = conn.cursor()
    # 校验账号密码及有效期
    cursor.execute("""
        SELECT user_id, expiry_date FROM users 
        WHERE username = ? AND password = ? AND is_active = 1
    """, (username, password))
    user = cursor.fetchone()

    if user:
        u_id, expiry_str = user
        expiry_date = datetime.datetime.strptime(expiry_str, '%Y-%m-%d').date()
        if expiry_date >= datetime.date.today():
            # 生成本次登录唯一的 Session ID
            new_sid = str(uuid.uuid4())
            cursor.execute("UPDATE users SET last_session_id = ? WHERE user_id = ?", (new_sid, u_id))
            conn.commit()
            conn.close()
            return True, u_id, expiry_str, new_sid
    conn.close()
    return False, None, None, None

# 设置路径
root_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if root_path not in sys.path:
    sys.path.insert(0, root_path)

try:
    from core.trade import start_optimization_task, VERSION
    from core.listing_logic import process_mercado_listing, get_column_options
except ImportError as e:
    st.error(f"核心模块导入失败: {e}")
    st.info(f"当前搜索路径: {sys.path[0]}")
    st.stop()

st.set_page_config(page_title=f"跨境AI大师 {VERSION}", layout="wide", page_icon="🚀")

# --- 核心状态初始化 ---
if 'optimization_done' not in st.session_state: st.session_state.optimization_done = False
if 'final_results' not in st.session_state: st.session_state.final_results = []
if 'process_logs' not in st.session_state: st.session_state.process_logs = []
if 'is_running' not in st.session_state: st.session_state.is_running = False
if 'current_df' not in st.session_state: st.session_state.current_df = None


# --- 停止按钮的回调函数 ---
def stop_optimization():
    st.session_state.is_running = False
    st.toast("正在请求停止任务...", icon="🛑")


# ================= 权限校验 (SaaS 账号版) =================
if 'authenticated' not in st.session_state: st.session_state.authenticated = False
if 'user_id' not in st.session_state: st.session_state.user_id = None
if 'session_id' not in st.session_state: st.session_state.session_id = None
if not st.session_state.authenticated:
    st.title("🔐 跨境AI大师 - 成员登录")
    col1, col2 = st.columns(2)
    with col1:
        u_name = st.text_input("账号")
    with col2:
        u_pwd = st.text_input("密码", type="password")

    if st.button("立即登录", use_container_width=True):
        is_ok, uid, expiry, sid = login_user(u_name, u_pwd)
        if is_ok:
            st.session_state.authenticated = True
            st.session_state.user_id = uid
            st.session_state.username = u_name
            st.session_state.session_id = sid
            st.session_state.expiry_date = expiry
            st.success(f"欢迎回来 {u_name}！权限有效期至: {expiry}")
            st.rerun()
        else:
            st.error("❌ 账号密码错误、已被禁用或已过期。")
    st.info("💡 如需开通账号，请联系管理员。")
    st.stop()

# --- 顶部状态条 ---
st.caption(f"👤 当前用户: {st.session_state.username} | 📅 到期时间: {st.session_state.expiry_date} | 🔑 状态: 独占登录中")

# ================= 侧边栏 =================
with st.sidebar:
    st.header("⚙️ 引擎配置")

    # --- 新增：优化模式选择 ---
    st.subheader("🎯 优化模式")
    opt_mode = st.radio(
        "选择模式",
        ["AI优化标题", "列组合优化"],
        help="【AI优化标题】: 仅对原标题进行SEO优化。\n【列组合优化】: 将选定列(如SKU)的内容强制作为关键词整合进标题。"
    )

    st.divider()
    st.subheader("🚫 违禁词过滤")
    negative_keywords = st.text_input(
        "输入禁止出现的词汇（多个请用逗号隔开）",
        placeholder="例如: Best, Cheap, Nike, Medical",
        help="AI在优化标题时会严格避开这些词汇"
    )

    # 🎯 核心新增：控制违禁词是拦截还是整行删除的开关
    delete_negative_rows = st.toggle(
        "💡 包含违禁词直接剔除整行",
        value=False,
        help="【关闭】: 走原来逻辑，AI 优化标题时会尝试避开或过滤这些词。\n【开启】: 前置扫描，只要原标题里含有以上任意一个违禁词，直接整行数据彻底删除，不消耗 AI 额度。"
    )

    engine_type = st.selectbox("AI 引擎", ["Google Gemini", "DeepSeek", "OpenAI GPT"])

    # 2. 根据引擎动态计算默认 URL
    if engine_type == "Google Gemini":
        default_url = "https://generativelanguage.googleapis.com/v1beta/openai/"
    elif engine_type == "DeepSeek":
        default_url = "https://api.deepseek.com"
    else:  # OpenAI GPT
        default_url = "https://api.openai.com/v1"

    base_url = st.text_input("API URL", value=default_url)

    # 3. 根据引擎动态切换模型选择
    if engine_type == "Google Gemini":
        gemini_opts = ["gemini-2.5-flash", "gemini-2.0-flash", "gemini-1.5-flash", "gemini-1.5-pro"]
        sel_m = st.selectbox("模型切换", gemini_opts, index=0)
        model_name = st.text_input("当前模型名称", value=sel_m)

    elif engine_type == "DeepSeek":
        model_opts = ["deepseek-chat", "deepseek-reasoner"]
        sel_m = st.selectbox("模型切换", model_opts, index=0)
        model_name = st.text_input("当前模型名称", value=sel_m)

    else:  # OpenAI GPT
        gpt_opts = ["gpt-4o", "gpt-4o-mini", "o1-mini", "gpt-4-turbo"]
        sel_m = st.selectbox("模型切换", gpt_opts, index=0)
        model_name = st.text_input("当前模型名称", value=sel_m)

    # 4. Key 接收区域
    raw_keys = st.text_area("🔑 Keys (每行一个)", height=100)
    user_keys = [k.strip() for k in raw_keys.split('\n') if k.strip()]

    st.divider()

    st.subheader("🛠️ 任务逻辑控制")
    use_deduplicate = st.checkbox("开启全局去重", value=True)
    deduplicate_limit = 999999
    if use_deduplicate:
        deduplicate_limit = st.number_input("分组去重上限", min_value=1, max_value=999, value=99)

    user_temperature = st.slider("🎨 AI 创造力", 0.0, 1.5, 0.7, 0.1)
    char_limit = st.slider("标题字符上限", 10, 200, 60)
    sleep_time = st.slider("🔥 间隔休眠时间 (秒)", 0.0, 50.0, 10.0, step=0.5)

    target_platform = st.selectbox("目标平台", ["Mercado Libre", "Amazon", "Shopee", "Rakuten.fr", "noon", "allegro"])
    target_lang = st.selectbox("目标语言", ["英语", "法语", "西班牙语", "葡萄牙语", "中文", "波兰语"])

    batch_size = st.number_input("📦 每批次处理个数", min_value=1, max_value=100, value=50)

# ================= 主界面 =================
tab_seo, tab_listing = st.tabs(["🔥 标题批量优化轮", "📦 美克多列表自动上架"])

with tab_seo:
    st.title("🚀 AI 标题优化引擎")
    selected_sheet = None
    selected_extra_cols = []
    # --- 新增/优化：一键清理按钮 ---
    # 我们使用一个大红色的按钮，并增加二次确认逻辑（或者直接重置）
    if st.session_state.optimization_done or st.session_state.current_df is not None or st.session_state.process_logs:
        if st.button("🗑️ 一键清空任务、日志与断点", use_container_width=True, type="secondary"):
            # 1. 重置任务状态
            st.session_state.optimization_done = False
            st.session_state.is_running = False

            # 2. 清空核心数据（断点和结果）
            st.session_state.current_df = None  # 清除断点备份的 DataFrame
            st.session_state.final_results = []  # 清除待下载的结果列表

            # 3. 清空界面反馈
            st.session_state.process_logs = []  # 清除运行日志

            # 4. 强制清理缓存（可选，防止上传组件残留文件名）
            st.cache_resource.clear()

            # 提示并刷新页面
            st.toast("已彻底清理所有任务数据", icon="🧹")
            time.sleep(0.5)
            st.rerun()
    if st.session_state.optimization_done:
        if st.button("🗑️ 清空所有结果与日志", use_container_width=True):
            st.session_state.optimization_done = False
            st.session_state.is_running = False
            st.session_state.final_results = []
            st.session_state.process_logs = []
            st.session_state.current_df = None
            st.cache_resource.clear()
            st.rerun()

    uploaded_files = st.file_uploader("上传待处理文件", type=['xlsx', 'csv'], accept_multiple_files=True)

    # --- 布局逻辑：导入文件后，分为左右两栏 ---
    if uploaded_files and not st.session_state.optimization_done:
        col_main, col_side_options = st.columns([2, 1])  # 左侧占2/3，右侧占1/3

        # --- 右侧：配置区 (Sheet 选择 + 表头选择) ---
        selected_extra_cols = []
        with col_side_options:
            if opt_mode == "列组合优化":
                st.subheader("🔗 关键词组合配置")
                try:
                    f = uploaded_files[0]
                    all_cols = []

                    # 1. 如果是 Excel，需要先选 Sheet
                    if f.name.endswith(('.xlsx', '.xls')):
                        excel_file = pd.ExcelFile(f)
                        sheet_names = excel_file.sheet_names
                        selected_sheet = st.selectbox("第一步：选择工作表 (Sheet)", options=sheet_names)

                        if selected_sheet:
                            # 2. 读取选中 Sheet 的表头
                            df_cols = pd.read_excel(f, sheet_name=selected_sheet, nrows=0)
                            all_cols = df_cols.columns.tolist()
                    else:
                        # CSV 文件直接读取表头
                        df_cols = pd.read_csv(f, nrows=0)
                        all_cols = df_cols.columns.tolist()

                    # 3. 展示多选框
                    if all_cols:
                        st.write("---")
                        selected_extra_cols = st.multiselect(
                            "第二步：勾选要组合的列 (如: SKU)",
                            options=all_cols,
                            help="选中的列内容将强制出现在AI生成的标题中",
                            placeholder="点击选择字段"
                        )

                    if not selected_extra_cols:
                        st.warning("⚠️ 请完成字段勾选")
                    else:
                        st.success(f"已选中 {len(selected_extra_cols)} 个字段")

                except Exception as e:
                    st.error(f"读取文件失败: {e}")
            else:
                st.info("💡 当前为标准优化模式\n无需配置额外字段")

        # --- 左侧：核心任务控制区 (逻辑保持不变) ---
        with col_main:
            if st.session_state.current_df is not None:
                st.warning("📊 检测到断点，将从上次位置继续")

            col_start, col_stop = st.columns(2)
            with col_start:
                btn_label = "▶️ 继续优化任务" if st.session_state.current_df is not None else "🔥 启动优化任务"
                start_btn = st.button(btn_label, type="primary", use_container_width=True,
                                      disabled=st.session_state.is_running)
            with col_stop:
                st.button("🛑 停止优化", use_container_width=True, on_click=stop_optimization,
                          disabled=not st.session_state.is_running)

            if start_btn:
                if not user_keys:
                    st.error("❌ 请先配置 API Key")
                elif opt_mode == "列组合优化" and not selected_extra_cols:
                    st.error("❌ 请先选择需要组合的关键词列")
                else:
                    st.session_state.is_running = True
                    if st.session_state.current_df is None:
                        st.session_state.process_logs = []
                    st.rerun()

            # 运行状态展示
            if st.session_state.is_running:
                with st.status("🚀 AI 正在努力工作中...", expanded=True) as status:
                    log_area = st.empty()

                    task_gen = start_optimization_task(
                        uploaded_files=uploaded_files,
                        platform=target_platform,
                        char_limit=char_limit,
                        language=target_lang,
                        api_keys=user_keys,
                        batch_size=batch_size,
                        sleep_time=sleep_time,
                        model_name=model_name,
                        base_url=base_url,
                        user_id=st.session_state.user_id,  # <--- 关键：用户隔离
                        current_sid=st.session_state.session_id,  # <--- 关键：防多端登录
                        use_deduplicate=use_deduplicate,
                        deduplicate_limit=deduplicate_limit,
                        temperature=user_temperature,
                        existing_df=st.session_state.current_df,
                        opt_mode=opt_mode,
                        selected_extra_cols=selected_extra_cols,
                        negative_keywords=negative_keywords,
                        selected_sheet=selected_sheet,  # 👈 确保这里传了 UI 右侧选中的值
                        delete_negative_rows=delete_negative_rows  # 🎯 注入灵魂：无缝打通前端控制开关
                    )

                    for msg in task_gen:
                        if not st.session_state.is_running:
                            break
                        if isinstance(msg, pd.DataFrame):
                            st.session_state.current_df = msg
                        elif isinstance(msg, list):
                            st.session_state.final_results = msg
                            st.session_state.optimization_done = True
                            st.session_state.is_running = False
                            status.update(label="✅ 任务圆满完成！", state="complete")
                        elif msg == "FINISH_SIGNAL":
                            pass
                        else:
                            l_entry = f"[{time.strftime('%H:%M:%S')}] {msg}"
                            st.session_state.process_logs.append(l_entry)
                            log_area.code("\n".join(st.session_state.process_logs[-15:]), language="bash")

                    st.rerun()

    # --- 处理完成后的下载展示区 ---
    if st.session_state.optimization_done:
        st.divider()  # 加一条分割线，视觉上更清晰
        st.subheader("✅ 优化结果清单预览")

        # 日志区增加“自动滚动到最新”感官
        with st.expander("📝 查看本次执行完整日志", expanded=False):
            st.code("\n".join(st.session_state.process_logs), language="bash")

        # 结果展示
        if not st.session_state.final_results:
            st.info("暂无处理结果。")
        else:
            st.markdown(f"**共计处理完成 {len(st.session_state.final_results)} 个文件**")
            d_cols = st.columns(3)

            for idx, (f_name, df_res) in enumerate(st.session_state.final_results):
                with d_cols[idx % 3]:
                    # 使用 os.path 提取文件名，比 rsplit 更稳健
                    base_name, ext = os.path.splitext(f_name)
                    is_csv = ext.lower() == '.csv'

                    try:
                        if is_csv:
                            # CSV 处理：utf-8-sig 是为了让 Excel 打开时不乱码（特别是西语特殊字符）
                            download_data = df_res.to_csv(index=False, encoding='utf-8-sig')
                            mime_type = "text/csv"
                            final_name = f"Opt_{base_name}.csv"
                        else:
                            # Excel 处理
                            buffer = io.BytesIO()
                            with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
                                df_res.to_excel(writer, index=False, sheet_name='Optimized')
                            download_data = buffer.getvalue()
                            mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            final_name = f"Opt_{base_name}.xlsx"

                        st.download_button(
                            label=f"📥 下载 {f_name}",
                            data=download_data,
                            file_name=final_name,
                            mime=mime_type,
                            key=f"dl_{idx}",
                            use_container_width=True
                        )
                    except Exception as e:
                        st.error(f"构建下载文件失败: {base_name}")

# ================= 这里是集成位置 =================

# 1. 在侧边栏定义超级管理员的入口
with st.sidebar:
    st.divider()
    with st.expander("🔐 管理员认证"):
        admin_token = st.text_input("超级管理员密钥", type="password", key="admin_key")
        # 建议把这里的密码改成你自己的私密字符串
        is_admin = (admin_token == "1194055104")

# 2. 连接数据库并显示管理面板 (放在页面最底部)
db_conn = sqlite3.connect(db_path, check_same_thread=False)
# 开启 WAL 模式确保管理操作不锁死主任务
db_conn.execute("PRAGMA journal_mode=WAL;")

manage_admin_and_history(
    db_conn,
    st.session_state.user_id,
    is_admin_mode=is_admin
)
db_conn.close()
# --- TAB 2: Listing (同步 Mexico 命名与逻辑定位) ---
with tab_listing:
    st.title("📦 自动化填充 (JSON+UPC)")

    json_conf_file = st.file_uploader("导入类目 JSON 配置", type=['json'])
    preset_vals = {}
    if json_conf_file:
        try:
            preset_vals = json.loads(json_conf_file.read().decode('utf-8-sig'))
            st.success("✅ JSON 加载成功")
        except:
            st.error("❌ JSON 格式错误")

    st.divider()
    c1, c2 = st.columns(2)
    with c1:
        src_file = st.file_uploader("上传结果表", type=['xlsx', 'csv', 'xls'], key="l_src")
        data_count = 0
        if src_file:
            try:
                # 1. 尝试读取逻辑
                src_file.seek(0)
                file_name = src_file.name.lower()
                src_df = None

                # 情况 A: 后缀是 .xlsx (但可能是伪装的 CSV)
                if file_name.endswith('.xlsx'):
                    try:
                        # 先按标准 Excel 读取
                        src_df = pd.read_excel(src_file, engine='openpyxl')
                    except Exception:
                        # 如果报错，说明可能是“手动改后缀”的 CSV，重置指针尝试 CSV 读取
                        src_file.seek(0)
                        try:
                            src_df = pd.read_csv(src_file, encoding='utf-8-sig')
                        except UnicodeDecodeError:
                            src_file.seek(0)
                            src_df = pd.read_csv(src_file, encoding='gbk')

                # 情况 B: 后缀是 .csv
                elif file_name.endswith('.csv'):
                    try:
                        src_df = pd.read_csv(src_file, encoding='utf-8-sig')
                    except UnicodeDecodeError:
                        src_file.seek(0)
                        src_df = pd.read_csv(src_file, encoding='gbk')

                # 情况 C: 后缀是 .xls
                elif file_name.endswith('.xls'):
                    try:
                        src_df = pd.read_excel(src_file, engine='xlrd')
                    except:
                        src_file.seek(0)
                        src_df = pd.read_html(src_file)[0]

                # 2. 校验结果并渲染 UI
                if src_df is not None and not src_df.empty:
                    # 💡 自动清理列名中的前后空格（防止匹配失败）
                    src_df.columns = [str(c).strip() for c in src_df.columns]

                    data_count = len(src_df)
                    st.info(f"📈 表格已加载：**{data_count}** 行数据")

                    t_col = st.selectbox("选择标题列", src_df.columns, key="sel_t_col_v3")
                    i_col = st.selectbox("选择图片列", src_df.columns, key="sel_i_col_v3")
                    s_col = st.selectbox("选择商品ID列 (用于生成SKU)", src_df.columns, key="sel_s_col_v3")

                    st.divider()
                    st.subheader("🔢 UPC 计数器")
                    upc_raw = st.text_area("在此粘贴 UPC (一行一个)", height=150, key="upc_area_v3")
                    u_list = [u.strip() for u in upc_raw.split('\n') if u.strip()]

                    if u_list:
                        upc_count = len(u_list)
                        if upc_count < data_count:
                            st.warning(f"⚠️ UPC 不足！还差 {data_count - upc_count} 个")
                        else:
                            st.success(f"✅ UPC 充足 (共 {upc_count} 个)")
                else:
                    st.error("❌ 读取到的表格数据为空，请检查文件内容。")

            except Exception as e:
                st.error(f"❌ 文件解析失败: {str(e)}")
                st.stop()
    with c2:
        tpl_file = st.file_uploader("上传美克多模板", type=['xlsx'], key="l_tpl")

        # --- 状态初始化 (用于持久化) ---
        if "gen_xlsx" not in st.session_state:
            st.session_state.gen_xlsx = None
        if "gen_json" not in st.session_state:
            st.session_state.gen_json = None
        if "gen_success" not in st.session_state:
            st.session_state.gen_success = False

        if tpl_file and src_file:
            tpl_bytes = tpl_file.getvalue()
            tpl_wb = openpyxl.load_workbook(io.BytesIO(tpl_bytes))
            target_sheet = st.selectbox("选择 Sheet", tpl_wb.sheetnames, index=len(tpl_wb.sheetnames) - 1)
            ws = tpl_wb[target_sheet]

            # --- 1. 核心识别逻辑 (原封不动) ---
            ml_headers, h_counts, header_row_idx = [], {}, None
            start_data_row = None
            char_col_idx = 2

            for r in range(1, 15):
                row_v_str = [str(ws.cell(row=r, column=c).value or "").replace('\n', ' ').replace('\r', ' ').strip() for
                             c in range(1, 100)]
                if any("Title" in x for x in row_v_str):
                    header_row_idx = r
                    for i, val in enumerate(row_v_str):
                        if "Number of characters" in val:
                            char_col_idx = i + 1
                            break

                    for v in row_v_str:
                        n = v if v != "None" and v != "" else ""
                        if n == "Mexico":
                            h_counts["Mexico"] = h_counts.get("Mexico", 0) + 1
                            if h_counts["Mexico"] == 1:
                                final_name = "Mexico (full)"
                            elif h_counts["Mexico"] == 2:
                                final_name = "Mexico"
                            else:
                                final_name = f"Mexico_{h_counts['Mexico']}"
                        else:
                            h_counts[n] = h_counts.get(n, 0) + 1
                            final_name = f"{n}_{h_counts[n]}" if h_counts[n] > 1 else n
                        ml_headers.append(final_name)
                    break

            if header_row_idx:
                for r in range(header_row_idx + 1, 40):
                    val_raw = ws.cell(row=r, column=char_col_idx).value
                    val_str = str(val_raw or "").strip()
                    if "=LEN" in val_str.upper():
                        start_data_row = r
                        break
                if not start_data_row: start_data_row = header_row_idx + 4

            st.success(f"📍 识别结果：字符限制在第 {char_col_idx} 列，起始行锁定为第 **{start_data_row}** 行")


            # --- 2. 过滤与匹配算法 (已修正 English 逻辑) ---
            def super_clean(text):
                if not text: return ""
                t = str(text).lower()
                t = t.replace("(full)", "")  # 不再删掉 english
                t = re.sub(r'\(.*?\)', '', t)
                t = re.sub(r'[^a-z0-9]', '', t)
                return t


            def should_show_header(h_name):
                h_low = h_name.lower()
                ch = super_clean(h_name)
                core_units = ["length", "width", "height", "package", "weight", "depth", "volume"]
                if "unit" in h_low and any(k in h_low for k in core_units):
                    return True
                if "mexico" in h_low or "description" in h_low:  # 显式放行
                    return True
                for jk in preset_vals.keys():
                    cjk = super_clean(jk)
                    if cjk == ch: return True
                    if len(cjk) > 5 and (cjk in ch or ch in cjk):
                        if "package" not in h_low and "package" in jk.lower():
                            if any(core in h_low for core in ["weight", "length", "width", "height"]):
                                continue
                        return True
                if re.search(r'_\d+$', h_name): return False
                # 垃圾词过滤不再包含 english
                garbage = ["select a value", "none", "inform product"]
                if any(g in h_low for g in garbage): return False
                return True


            clean_h = [h for h in ml_headers if h and h != "" and should_show_header(h)]

            # --- 3. 自动勾选逻辑 (原封不动) ---
            auto_sel = []
            for h in clean_h:
                if "Title:" in h and "inform product" in h: continue
                if "(full)" in h.lower(): continue
                h_low, ch = h.lower(), super_clean(h)
                is_matched = False
                if "unit" in h_low and any(k in h_low for k in ["length", "width", "height", "weight", "volume"]):
                    for jk, jv in preset_vals.items():
                        jk_low = jk.lower()
                        if "unit" in jk_low:
                            if ("package" in h_low) == ("package" in jk_low):
                                if any(k in h_low and k in jk_low for k in
                                       ["length", "width", "height", "weight", "volume"]):
                                    auto_sel.append(h)
                                    is_matched = True
                                    break
                if not is_matched:
                    for jk in preset_vals.keys():
                        if super_clean(jk) == ch:
                            auto_sel.append(h)
                            is_matched = True
                            break
                    if not is_matched:
                        for jk in preset_vals.keys():
                            cjk = super_clean(jk)
                            if (len(cjk) > 3 and (cjk in ch or ch in cjk)) or (
                                    "description" in cjk and "description" in ch):
                                if "package" in jk.lower() and "package" not in h_low:
                                    if any(core in h_low for core in ["weight", "length", "width", "height"]):
                                        continue
                                auto_sel.append(h)
                                break

            auto_sel = list(dict.fromkeys(auto_sel))
            to_fill = st.multiselect("确认填充属性：", clean_h, default=auto_sel)

            # --- 4. 静态填充渲染 (修复：使用属性名作为唯一 Key) ---
            static_data = {}
            for i, h in enumerate(to_fill):
                h_low, ch = h.lower(), super_clean(h)
                matched_val = None

                for jk, jv in preset_vals.items():
                    if super_clean(jk) == ch: matched_val = jv; break
                if matched_val is None and "unit" in h_low:
                    for jk, jv in preset_vals.items():
                        jk_low = jk.lower()
                        if "unit" in jk_low:
                            if ("package" in h_low) == ("package" in jk_low):
                                if any(k in h_low and k in jk_low for k in
                                       ["length", "width", "height", "weight", "volume"]):
                                    matched_val = jv;
                                    break
                if matched_val is None:
                    for jk, jv in preset_vals.items():
                        cjk = super_clean(jk)
                        if "unit" in h_low and "unit" not in jk.lower(): continue
                        if (len(cjk) > 3 and (cjk in ch or ch in cjk)) or (
                                "description" in cjk and "description" in ch):
                            if "package" in jk.lower() and "package" not in h_low:
                                if any(core in h_low for core in ["weight", "length", "width", "height"]):
                                    continue
                            matched_val = jv;
                            break

                col_idx = ml_headers.index(h) + 1
                opts = get_column_options(ws, tpl_wb, col_idx, header_row_idx)
                val_str = str(matched_val) if matched_val is not None else ""

                # --- 使用属性名 h 作为唯一 Key，防止删除行时数据覆盖 ---
                if "description" in h_low or "description" in ch:
                    static_data[h] = st.text_area(f"[{h}]", value=val_str, key=f"area_{h}", height=180)
                elif opts:
                    try:
                        d_idx = opts.index(val_str)
                        static_data[h] = st.selectbox(f"[{h}]", opts, index=d_idx, key=f"sel_{h}")
                    except:
                        static_data[h] = st.text_input(f"[{h}]", value=val_str, key=f"in_m_{h}")
                else:
                    static_data[h] = st.text_input(f"[{h}]", value=val_str, key=f"txt_{h}")

            # --- 5. 生成与清除按钮区域 ---
            st.divider()
            col_b1, col_b2 = st.columns(2)
            with col_b1:
                if st.button("🚀 开始生成上架表", type="primary", use_container_width=True, key="main_run_btn"):
                    if not static_data:
                        st.warning("⚠️ 请选择属性")
                    else:
                        with st.spinner("生成中..."):
                            res_xlsx, _ = process_mercado_listing(
                                source_df=src_df,
                                template_bytes=tpl_bytes,
                                sheet_name=target_sheet,
                                mapping_config={'title_col': t_col, 'img_col': i_col, 'sku_col': s_col},
                                static_fills=static_data,
                                upc_list=u_list,
                                start_row=start_data_row
                            )
                            # 生成干净的 JSON（保留星号/括号，去索引）
                            clean_json = {re.sub(r'_\d+$', '', k): v for k, v in static_data.items()}

                            st.session_state.gen_xlsx = res_xlsx
                            st.session_state.gen_json = json.dumps(clean_json, ensure_ascii=False, indent=4)
                            st.session_state.gen_success = True

            with col_b2:
                if st.button("🗑️ 清除生成结果", use_container_width=True, key="clear_state_btn"):
                    st.session_state.gen_xlsx = None
                    st.session_state.gen_json = None
                    st.session_state.gen_success = False
                    st.rerun()

            # --- 6. 结果下载区域 ---
            if st.session_state.gen_success and st.session_state.gen_xlsx:
                st.success("✅ 文件已锁定，可多次下载。")
                d1, d2 = st.columns(2)
                import time

                ts = int(time.time())
                with d1:
                    st.download_button("📥 下载 Excel", data=st.session_state.gen_xlsx, file_name=f"Ready_{ts}.xlsx",
                                       key="dl_x_final")
                with d2:
                    st.download_button("📄 下载 JSON", data=st.session_state.gen_json, file_name=f"Config_{ts}.json",
                                       key="dl_j_final")