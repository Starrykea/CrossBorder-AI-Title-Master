import openpyxl
from openpyxl.cell import MergedCell
from openpyxl.utils import range_boundaries, get_column_letter
from openpyxl.styles import Alignment
import io
import time
import random
import re
import json

def get_column_options(ws, tpl_workbook, col_idx, header_row_idx):
    """从 Excel 模板中动态解析下拉选项 (用于 UI 渲染)"""
    # 假设美克多的有效数据从表头下第 4 行开始，我们检查那一行的单元格是否有数据验证
    target_cell_coord = f"{get_column_letter(col_idx)}{header_row_idx + 4}"
    options = []
    header_name = str(ws.cell(row=header_row_idx, column=col_idx).value or "")

    # 1. 常见跨境电商字段保底选项
    if "Warranty type" in header_name:
        return ["No warranty", "Seller warranty", "Factory warranty"]
    if "Brand" in header_name:
        return ["generic"]
    if "weight unit" in header_name.lower():
        return ["g", "kg", "lb", "mg", "oz"]
    if any(k in header_name.lower() for k in ["length", "width", "height"]):
        if "unit" in header_name.lower():
            return ["cm", "\"", "m", "mm", "ft"]

    # 2. 从 Excel 内部的数据验证 (Data Validation) 提取列表
    if hasattr(ws, 'data_validations') and ws.data_validations:
        for dv in ws.data_validations.dataValidation:
            if target_cell_coord in dv:
                formula = dv.formula1
                if not formula: continue
                # 如果是直接写在公式里的列表 (如 "Yes,No")
                if '"' in formula or ("," in formula and "!" not in formula):
                    options = formula.replace('"', '').split(',')
                # 如果是引用其他 Sheet 的序列 (如 =Lists!$A$1:$A$10)
                elif "!" in formula:
                    try:
                        sheet_part, range_part = formula.split('!')
                        sheet_name = sheet_part.strip("='")
                        if sheet_name in tpl_workbook.sheetnames:
                            ref_ws = tpl_workbook[sheet_name]
                            clean_range = range_part.replace("$", "")
                            min_col, min_row, max_col, max_row = range_boundaries(clean_range)
                            for r in range(min_row, max_row + 1):
                                val = ref_ws.cell(row=r, column=min_col).value
                                if val: options.append(str(val).strip())
                    except:
                        pass
                break
    return list(set(options))


def safe_write(ws, row, col, value):
    """安全写入，支持数字转换与自动换行"""
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell): return
    if value is None: return

    val_str = str(value).strip()
    # 针对 Description 等长文本开启自动换行
    if "\n" in val_str:
        cell.alignment = Alignment(wrapText=True, vertical='top')

    # 尝试将纯数字字符串转为真正的数字类型，避免 Excel 警告
    if val_str.isdigit():
        cell.value = int(val_str)
    else:
        try:
            cell.value = float(val_str)
        except:
            cell.value = val_str


def process_mercado_listing(source_df, template_bytes, sheet_name, mapping_config, static_fills, upc_list=None,
                            start_row=None):
    import json
    import io
    import re
    import openpyxl

    in_io = io.BytesIO(template_bytes)
    wb = openpyxl.load_workbook(in_io)
    ws = wb[sheet_name]

    # --- 1. 寻找表头逻辑 (后端：必须与 UI 逻辑完全同步) ---
    header_row_idx = 1
    headers = []
    h_counts = {}

    for i in range(1, 25):
        row_vals_raw = [str(ws.cell(row=i, column=j).value or "").strip() for j in range(1, 100)]
        if any("Title" in v for v in row_vals_raw):
            header_row_idx = i
            for v in row_vals_raw:
                n = v.replace('\n', ' ').strip()
                if n == "" or n == "None":
                    headers.append("")
                elif n == "Mexico":
                    h_counts["Mexico"] = h_counts.get("Mexico", 0) + 1
                    if h_counts["Mexico"] == 1:
                        final_name = "Mexico (full)"
                    else:
                        final_name = "Mexico"
                    headers.append(final_name)
                else:
                    h_counts[n] = h_counts.get(n, 0) + 1
                    headers.append(f"{n}_{h_counts[n]}" if h_counts[n] > 1 else n)
            break

    col_map = {name: idx + 1 for idx, name in enumerate(headers) if name != ''}

    # 锁定数据起始行
    char_count_col_idx = 2
    if start_row is None:
        for r in range(header_row_idx + 1, 60):
            val_str = str(ws.cell(row=r, column=char_count_col_idx).value or "").strip()
            if "=LEN" in val_str.upper():
                start_row = r
                break
        if not start_row: start_row = header_row_idx + 4

    # 功能列识别
    ml_upc_key = next((k for k in col_map.keys() if "Universal product code" in k), None)
    sku_key = next((k for k in col_map.keys() if "SKU" in k), None)
    all_color_keys = [k for k in col_map.keys() if "color" in k.lower()]

    # --- 修改部分：颜色递增配置初始化 (支持多列独立识别) ---
    color_inc_configs = {}
    for ck in all_color_keys:
        if ck in static_fills:
            val = str(static_fills[ck])
            # 尝试匹配 字母+数字 格式 (如 Aa101)
            color_match = re.match(r"([a-zA-Z]+)([0-9]+)", val)
            if color_match:
                color_inc_configs[ck] = {
                    "prefix": color_match.group(1),
                    "start_num": int(color_match.group(2))
                }
    # ----------------------------------------------------

    # --- 2. 核心填充循环 ---
    for i, (_, row_data) in enumerate(source_df.iterrows()):
        curr_row = start_row + i

        # A. 基础字段
        real_title_key = next((k for k in col_map.keys() if "Title" in k and "Number" not in k), None)
        if real_title_key:
            title_v = str(row_data[mapping_config['title_col']])
            if len(title_v) > 60:
                title_v = title_v.replace("Comfortable", "Comfort").replace("Artwork", "Art").replace("Car ", "")
                title_v = title_v[:60].strip()
            safe_write(ws, curr_row, col_map[real_title_key], title_v)

        photo_key = next((k for k in col_map.keys() if "Photos" in k), None)
        if photo_key: safe_write(ws, curr_row, col_map[photo_key], row_data[mapping_config['img_col']])

        if sku_key and mapping_config.get('sku_col'):
            sku_v = f"{str(row_data[mapping_config['sku_col']]).strip()}-M{i + 1}"
            safe_write(ws, curr_row, col_map[sku_key], sku_v)

        if ml_upc_key and upc_list and i < len(upc_list):
            safe_write(ws, curr_row, col_map[ml_upc_key], upc_list[i])

        # B. 静态配置字段填充
        for header_key, val in static_fills.items():
            if header_key in col_map:
                target_col = col_map[header_key]
                h_low = header_key.lower()
                final_val = val

                # --- 修改部分：根据列名精准判断是否递增 ---
                if header_key in color_inc_configs:
                    cfg = color_inc_configs[header_key]
                    final_val = f"{cfg['prefix']}{cfg['start_num'] + i}"
                # ----------------------------------------

                elif "description" in h_low:
                    clean_v = re.sub(r'</p>|<br\s*/?>', '\n', str(val))
                    clean_v = re.sub(r'<[^>]+>', '', clean_v)
                    final_val = re.sub(r'\n\s*\n', '\n', clean_v).strip()

                safe_write(ws, curr_row, target_col, final_val)

    out_io = io.BytesIO()
    wb.save(out_io)
    return out_io.getvalue(), ""